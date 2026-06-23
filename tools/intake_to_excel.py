#!/usr/bin/env python3
"""
intake_to_excel.py — render the canonical intake JSON (from project_intake.py)
into a presentation `.xlsx` workbook (phase 2 of /project-intake).

PURE rendering: no analysis logic lives here. Every value comes straight from the
canonical document, so the workbook and the JSON can never disagree. Each sheet
carries a *Provenance* column (static / tmc / manual) and is colour-coded, so a
reader sees at a glance what was derived from code vs. what still needs TMC or a
human.

openpyxl is an OPTIONAL dependency (only this renderer needs it; the phase-1
analyzer stays stdlib-only so the canonical JSON is always producible). Install
with `pip install openpyxl` — see INSTALL.md.

Usage:
    intake_to_excel.py --in intake.json --out intake.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent))
import talend_complexity as tcx  # noqa: E402  (stdlib-only; used for the signal-key list)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover - exercised only when dependency absent
    _HAVE_OPENPYXL = False

_HEADER_FILL = "1F3864"      # dark blue
_PROV_FILL = {"static": "E2EFDA", "tmc": "DDEBF7", "manual": "FCE4D6"}  # green / blue / amber


def _require_openpyxl() -> None:
    if not _HAVE_OPENPYXL:
        raise RuntimeError(
            "openpyxl is required for the Excel renderer. Install it with "
            "`pip install openpyxl` (the phase-1 analyzer needs no dependencies)."
        )


def _names(doc: dict) -> dict[str, str]:
    return {a["artifact_id"]: a["name"] for a in doc["artifacts"]}


def _sysmap(doc: dict) -> dict[str, dict]:
    return {s["system_id"]: s for s in doc["systems"]}


def _header(ws, headers: list[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _prov_fill(row_cells, provenance: str) -> None:
    color = _PROV_FILL.get(provenance)
    if color:
        for cell in row_cells:
            cell.fill = PatternFill("solid", fgColor=color)


# --------------------------------------------------------------------------- #
# Sheets
# --------------------------------------------------------------------------- #
def sheet_summary(wb, doc: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    p = doc["project"]
    ws["A1"] = "Talend Project Intake — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Project", p["name"]),
        ("Talend version", p["product_version"]),
        ("Scanned path", p["scanned_path"]),
        ("Generated at", doc["generated_at"]),
        ("Config version", doc["generator"].get("config_version", "")),
        ("", ""),
        ("Distinct external systems", len(doc["systems"])),
        ("Proposed interfaces", len(doc["interfaces"])),
        ("Gaps to resolve", len(doc["gaps"])),
        ("Parse errors", len(p.get("parse_errors", []))),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1

    ws.cell(row=r + 1, column=1, value="Artifact counts").font = Font(bold=True, size=12)
    r += 2
    for k, v in p["artifact_counts"].items():
        if v:
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            r += 1

    ws.cell(row=r + 1, column=1, value="Complexity (estimated, uncalibrated)").font = Font(bold=True, size=12)
    r += 2
    from collections import Counter
    hist = Counter(a["complexity"]["bucket"] for a in doc["artifacts"]
                   if a.get("complexity"))
    for b in ("Very Simple", "Simple", "Moderate", "Complex", "Very Complex"):
        if hist.get(b):
            ws.cell(row=r, column=1, value=b)
            ws.cell(row=r, column=2, value=hist[b])
            r += 1

    # Findings overview (one line per severity) — only when something was found.
    findings = doc.get("findings", {})
    by_sev = findings.get("by_severity", {})
    if findings.get("total"):
        ws.cell(row=r + 1, column=1, value="Findings (breadth pass)").font = Font(bold=True, size=12)
        r += 2
        order = sorted(by_sev.items(), key=lambda kv: _SEVERITY_ORDER.get(kv[0], 99))
        summary_line = ", ".join(f"{sev}: {n}" for sev, n in order)
        ws.cell(row=r, column=1, value="By severity")
        ws.cell(row=r, column=2, value=f"{findings['total']} total ({summary_line})")
        r += 1

    # Dependencies + upgrade-risk overview (only when something was detected).
    deps = doc.get("dependencies", {})
    if deps.get("distinct_jars") or deps.get("version_drift"):
        ws.cell(row=r + 1, column=1, value="Dependencies & upgrade risk").font = Font(bold=True, size=12)
        r += 2
        for label, val in (("Distinct jars", len(deps.get("distinct_jars", []))),
                           ("Distinct libraries", len(deps.get("distinct_libs", []))),
                           ("Total references", deps.get("total_refs", 0)),
                           ("Version-pinned jars", len(deps.get("version_pinned_jars", []))),
                           ("Libraries with version drift", len(deps.get("version_drift", [])))):
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        for flag in deps.get("upgrade_risk_flags", []):
            ws.cell(row=r, column=1, value="⚠ " + flag).font = Font(italic=True)
            r += 1

    # TMC overview (only when enriched).
    tmc = doc.get("tmc", {})
    if tmc.get("enriched"):
        summ = tmc.get("summary", {})
        ws.cell(row=r + 1, column=1, value="TMC (read-only)").font = Font(bold=True, size=12)
        r += 2
        by_env = summ.get("deployment_by_environment", {})
        for label, val in (("Region", tmc.get("region")),
                           ("Environments", len(doc.get("environments", []))),
                           ("Engines", len(doc.get("infrastructure", {}).get("engines", []))),
                           ("Plans", len(tmc.get("plans", []))),
                           ("Deployable total", summ.get("deployable_total")),
                           ("Deployed", summ.get("deployed")),
                           ("Deployed in prod", summ.get("deployed_in_prod")),
                           ("Reachable via parent (worker)", summ.get("reachable_via_parent")),
                           ("Orphaned candidates", len(summ.get("orphaned_candidates", []))),
                           ("Per environment", ", ".join(f"{e}: {n}" for e, n in sorted(by_env.items())))):
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1

    ws.cell(row=r + 1, column=1, value="Provenance legend").font = Font(bold=True, size=12)
    r += 2
    for prov, desc in (("static", "derived from the Talend project on disk"),
                       ("tmc", "from Talend Management Console (phase 3)"),
                       ("manual", "needs human/customer input (phase 4)")):
        c1 = ws.cell(row=r, column=1, value=prov)
        c1.fill = PatternFill("solid", fgColor=_PROV_FILL[prov])
        ws.cell(row=r, column=2, value=desc)
        r += 1
    _autosize(ws, [30, 70])


def sheet_infrastructure(wb, doc: dict) -> None:
    ws = wb.create_sheet("Infrastructure")
    infra = doc.get("infrastructure", {})
    envs = doc.get("environments", [])
    engines = infra.get("engines", [])
    workspaces = infra.get("workspaces", [])
    _header(ws, ["Category", "Name", "Detail", "Status", "Provenance"])
    r = 2
    if not envs and not engines and not workspaces:
        ws.cell(row=r, column=1, value="(empty — populate via TMC enrichment, phase 3)")
        _autosize(ws, [16, 28, 40, 16, 12])
        return

    region = infra.get("tmc_region")
    if region:
        cells = [ws.cell(row=r, column=1, value="TMC region"),
                 ws.cell(row=r, column=2, value=region),
                 ws.cell(row=r, column=3, value=""),
                 ws.cell(row=r, column=4, value=""),
                 ws.cell(row=r, column=5, value="tmc")]
        _prov_fill(cells, "tmc")
        r += 1

    for env in envs:
        detail_bits = []
        if env.get("max_cloud_containers") is not None:
            detail_bits.append(f"max cloud containers: {env['max_cloud_containers']}")
        if env.get("is_default"):
            detail_bits.append("default")
        prov = env.get("provenance", "tmc")
        cells = [ws.cell(row=r, column=1, value="Environment"),
                 ws.cell(row=r, column=2, value=env.get("name")),
                 ws.cell(row=r, column=3, value=env.get("description") or "; ".join(detail_bits)),
                 ws.cell(row=r, column=4, value="default" if env.get("is_default") else ""),
                 ws.cell(row=r, column=5, value=prov)]
        _prov_fill(cells, prov)
        r += 1

    for ws_def in workspaces:
        prov = ws_def.get("provenance", "tmc")
        detail = ", ".join(b for b in (ws_def.get("type"),
                                       f"env={ws_def['environment_id']}" if ws_def.get("environment_id") else None,
                                       f"owner={ws_def['owner']}" if ws_def.get("owner") else None) if b)
        cells = [ws.cell(row=r, column=1, value="Workspace"),
                 ws.cell(row=r, column=2, value=ws_def.get("name")),
                 ws.cell(row=r, column=3, value=detail),
                 ws.cell(row=r, column=4, value=""),
                 ws.cell(row=r, column=5, value=prov)]
        _prov_fill(cells, prov)
        r += 1

    for eng in engines:
        prov = eng.get("provenance", "tmc")
        # Engine identity differs between static (name/run_profiles/status) and the
        # TMC enrichment (engine_id/package_version/n_services). Render whichever is present.
        name = eng.get("name") or eng.get("engine_id")
        detail_bits = []
        if eng.get("run_profiles"):
            detail_bits.append(",".join(eng["run_profiles"]))
        if eng.get("package_version"):
            detail_bits.append(f"package: {eng['package_version']}")
        if eng.get("n_services") is not None:
            detail_bits.append(f"{eng['n_services']} service(s)")
        cells = [ws.cell(row=r, column=1, value="Engine"),
                 ws.cell(row=r, column=2, value=name),
                 ws.cell(row=r, column=3, value="; ".join(detail_bits)),
                 ws.cell(row=r, column=4, value=eng.get("status")),
                 ws.cell(row=r, column=5, value=prov)]
        _prov_fill(cells, prov)
        r += 1

    for note in infra.get("manual_notes", []):
        cells = [ws.cell(row=r, column=1, value="Manual note"),
                 ws.cell(row=r, column=2, value=""),
                 ws.cell(row=r, column=3, value=note.get("text")),
                 ws.cell(row=r, column=4, value=""),
                 ws.cell(row=r, column=5, value="manual")]
        _prov_fill(cells, "manual")
        r += 1
    _autosize(ws, [16, 28, 44, 16, 12])


def sheet_artifacts(wb, doc: dict) -> None:
    ws = wb.create_sheet("Artifacts")
    _header(ws, ["Name", "Type", "Version", "Folder", "Complexity", "Score",
                 "#Read", "#Write", "Calls", "Flags", "Provenance"])
    r = 2
    for a in sorted(doc["artifacts"], key=lambda x: (x["type"], x["name"])):
        cx = a.get("complexity") or {}
        cells = [
            ws.cell(row=r, column=1, value=a["name"]),
            ws.cell(row=r, column=2, value=a["type"]),
            ws.cell(row=r, column=3, value=a["item_version"]),
            ws.cell(row=r, column=4, value=str(Path(a["path"]).parent)),
            ws.cell(row=r, column=5, value=cx.get("bucket")),
            ws.cell(row=r, column=6, value=cx.get("score")),
            ws.cell(row=r, column=7, value=len(a["systems_read"])),
            ws.cell(row=r, column=8, value=len(a["systems_write"])),
            ws.cell(row=r, column=9, value=", ".join(c["target_name"] for c in a["calls"])),
            ws.cell(row=r, column=10, value="; ".join(a.get("non_standard_flags", []))),
            ws.cell(row=r, column=11, value=a["provenance"]),
        ]
        _prov_fill(cells, a["provenance"])
        r += 1
    ws.auto_filter.ref = f"A1:K{max(2, r - 1)}"
    _autosize(ws, [34, 12, 9, 28, 14, 8, 7, 7, 30, 30, 12])


def sheet_systems(wb, doc: dict) -> None:
    ws = wb.create_sheet("Systems")
    _header(ws, ["Family", "Technology", "Host", "Database", "Schema", "URI/Endpoint",
                 "Objects", "Resolved", "Confidence", "Provenance"])
    r = 2
    for s in doc["systems"]:
        ident = s["identity"]
        cells = [
            ws.cell(row=r, column=1, value=s["family"]),
            ws.cell(row=r, column=2, value=s["technology"]),
            ws.cell(row=r, column=3, value=ident.get("host")),
            ws.cell(row=r, column=4, value=ident.get("database")),
            ws.cell(row=r, column=5, value=ident.get("schema")),
            ws.cell(row=r, column=6, value=ident.get("uri") if ident.get("uri") != "(unresolved)"
                    else ident.get("endpoint")),
            ws.cell(row=r, column=7, value=", ".join(s.get("objects", [])[:10])),
            ws.cell(row=r, column=8, value="yes" if s["resolved"] else "no"),
            ws.cell(row=r, column=9, value=s["confidence"]),
            ws.cell(row=r, column=10, value=s["provenance"]),
        ]
        _prov_fill(cells, s["provenance"])
        r += 1
    ws.auto_filter.ref = f"A1:J{max(2, r - 1)}"
    _autosize(ws, [14, 22, 22, 18, 14, 30, 30, 9, 11, 12])


def sheet_system_usage(wb, doc: dict) -> None:
    """Read/write matrix, flattened to one row per system with the artifacts that touch it."""
    ws = wb.create_sheet("System Read-Write")
    _header(ws, ["Family", "Technology", "Read by", "Written by"])
    names = _names(doc)
    r = 2
    for s in doc["systems"]:
        sid = s["system_id"]
        read_by = [names[a["artifact_id"]] for a in doc["artifacts"] if sid in a["systems_read"]]
        write_by = [names[a["artifact_id"]] for a in doc["artifacts"] if sid in a["systems_write"]]
        ws.cell(row=r, column=1, value=s["family"])
        ws.cell(row=r, column=2, value=s["technology"])
        ws.cell(row=r, column=3, value=", ".join(sorted(read_by)))
        ws.cell(row=r, column=4, value=", ".join(sorted(write_by)))
        r += 1
    _autosize(ws, [14, 22, 50, 50])


def sheet_complexity(wb, doc: dict) -> None:
    ws = wb.create_sheet("Complexity")
    signal_keys = list(tcx.DEFAULT_CONFIG["signals"].keys())
    lead = ["Name", "Type", "Bucket", "Score", "Calibrated", "Needs LLM review", "LLM rating"]
    _header(ws, lead + signal_keys)
    r = 2
    for a in sorted(doc["artifacts"], key=lambda x: -(x.get("complexity") or {}).get("score", 0)):
        cx = a.get("complexity")
        if not cx:
            continue
        ws.cell(row=r, column=1, value=a["name"])
        ws.cell(row=r, column=2, value=a["type"])
        ws.cell(row=r, column=3, value=cx["bucket"])
        ws.cell(row=r, column=4, value=cx["score"])
        ws.cell(row=r, column=5, value="yes" if cx["calibrated"] else "no")
        ws.cell(row=r, column=6, value="yes" if cx.get("needs_llm_review") else "no")
        ws.cell(row=r, column=7, value=cx.get("llm_rating"))
        for i, k in enumerate(signal_keys, start=len(lead) + 1):
            ws.cell(row=r, column=i, value=cx["signals"].get(k))
        r += 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(lead) + len(signal_keys))}{max(2, r - 1)}"
    _autosize(ws, [34, 12, 14, 8, 10, 16, 11] + [10] * len(signal_keys))


_SEVERITY_ORDER = {"bug": 0, "perf": 1, "smell": 2, "dead_code": 3}


def sheet_findings(wb, doc: dict) -> None:
    """One row per finding, flattened from artifacts[].findings. Pure lookup:
    severity/category/detail/location/provenance come straight from each finding;
    the artifact name is looked up from its containing artifact. Sorted by
    severity (bug → perf → smell → dead_code) then category."""
    ws = wb.create_sheet("Findings")
    _header(ws, ["Artifact", "Severity", "Category", "Location", "Detail", "Provenance"])
    rows = []
    for a in doc["artifacts"]:
        for f in a.get("findings", []):
            rows.append((a["name"], f))
    rows.sort(key=lambda x: (_SEVERITY_ORDER.get(x[1].get("severity"), 99),
                             x[1].get("category") or "",
                             x[0]))
    r = 2
    if not rows:
        ws.cell(row=r, column=1, value="(no findings — deterministic breadth pass "
                                       "found nothing; see complexity.needs_llm_review "
                                       "for the depth pass)")
        _autosize(ws, [30, 12, 20, 24, 60, 12])
        return
    for name, f in rows:
        prov = f.get("provenance", "static")
        cells = [
            ws.cell(row=r, column=1, value=name),
            ws.cell(row=r, column=2, value=f.get("severity")),
            ws.cell(row=r, column=3, value=f.get("category")),
            ws.cell(row=r, column=4, value=f.get("location")),
            ws.cell(row=r, column=5, value=f.get("detail")),
            ws.cell(row=r, column=6, value=prov),
        ]
        _prov_fill(cells, prov)
        r += 1
    ws.auto_filter.ref = f"A1:F{max(2, r - 1)}"
    _autosize(ws, [30, 12, 20, 24, 60, 12])


def sheet_interfaces(wb, doc: dict) -> None:
    ws = wb.create_sheet("Interfaces")
    _header(ws, ["Interface", "Label", "Status", "Confidence", "Members",
                 "Entry points", "Systems touched", "Ambiguous", "Provenance"])
    names = _names(doc)
    sysmap = _sysmap(doc)
    r = 2
    for i in doc["interfaces"]:
        cells = [
            ws.cell(row=r, column=1, value=i["interface_id"]),
            ws.cell(row=r, column=2, value=i["label"]),
            ws.cell(row=r, column=3, value=i["status"]),
            ws.cell(row=r, column=4, value=i["confidence"]),
            ws.cell(row=r, column=5, value=", ".join(names.get(m, m) for m in i["member_artifacts"])),
            ws.cell(row=r, column=6, value=", ".join(names.get(m, m) for m in i["entry_points"])),
            ws.cell(row=r, column=7, value=", ".join(sysmap[s]["technology"]
                    for s in i["systems_touched"] if s in sysmap)),
            ws.cell(row=r, column=8, value=", ".join(names.get(m, m) for m in i["ambiguous_members"])),
            ws.cell(row=r, column=9, value=i["provenance"]),
        ]
        _prov_fill(cells, i["provenance"])
        r += 1
    _autosize(ws, [16, 24, 12, 12, 44, 28, 36, 24, 12])


def sheet_orchestration(wb, doc: dict) -> None:
    ws = wb.create_sheet("Orchestration (TMC)")
    _header(ws, ["Plan", "Description", "Workspace", "Environment",
                 "#Flows", "Flow IDs", "Provenance"])
    r = 2
    plans = doc.get("tmc", {}).get("plans", [])
    if not plans:
        ws.cell(row=r, column=1, value="(empty — populate via TMC enrichment, phase 3: "
                                       "plans tell you which jobs run together)")
        _autosize(ws, [22, 30, 14, 14, 8, 36, 12])
        return
    for plan in plans:
        prov = plan.get("provenance", "tmc")
        cells = [ws.cell(row=r, column=1, value=plan.get("name")),
                 ws.cell(row=r, column=2, value=plan.get("description")),
                 ws.cell(row=r, column=3, value=plan.get("workspace_id")),
                 ws.cell(row=r, column=4, value=plan.get("environment_id")),
                 ws.cell(row=r, column=5, value=plan.get("n_flows", len(plan.get("flow_ids", [])))),
                 ws.cell(row=r, column=6, value=", ".join(plan.get("flow_ids", []))),
                 ws.cell(row=r, column=7, value=prov)]
        _prov_fill(cells, prov)
        r += 1
    _autosize(ws, [22, 30, 14, 14, 8, 36, 12])


def sheet_deployment(wb, doc: dict) -> None:
    """Per-artifact deployed/worker/orphaned classification, driven by tmc.summary
    + artifacts[].tmc_task. Degrades to a placeholder on a static-only doc."""
    ws = wb.create_sheet("Deployment")
    tmc = doc.get("tmc", {})
    summary = tmc.get("summary", {})
    if not tmc.get("enriched") and not summary:
        _header(ws, ["Artifact", "Type", "Deployed?", "Environments", "In prod?",
                     "Classification", "Task IDs", "Provenance"])
        ws.cell(row=2, column=1, value="(empty — populate via TMC enrichment, phase 3: "
                                       "deployed-vs-source classification needs TMC tasks)")
        _autosize(ws, [30, 12, 11, 22, 9, 18, 20, 12])
        return

    # Overview block (counts from tmc.summary) first, then the per-artifact table.
    ws.cell(row=1, column=1, value="Deployment overview").font = Font(bold=True, size=12)
    orphaned = summary.get("orphaned_candidates", [])
    overview = [
        ("Deployable total", summary.get("deployable_total")),
        ("Deployed (own TMC task)", summary.get("deployed")),
        ("Deployed in prod", summary.get("deployed_in_prod")),
        ("Reachable via parent (worker)", summary.get("reachable_via_parent")),
        ("Orphaned candidates", len(orphaned)),
        ("Unmatched TMC tasks", len(summary.get("unmatched_tasks", []))),
    ]
    r = 2
    for label, val in overview:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    by_env = summary.get("deployment_by_environment", {})
    if by_env:
        ws.cell(row=r, column=1, value="Per environment").font = Font(bold=True)
        ws.cell(row=r, column=2, value=", ".join(f"{e}: {n}" for e, n in sorted(by_env.items())))
        r += 1
    note = summary.get("note")
    if note:
        ws.cell(row=r, column=1, value="Note").font = Font(bold=True)
        ws.cell(row=r, column=2, value=note).alignment = Alignment(wrap_text=True)
        r += 1

    # Per-artifact table. Classification mirrors tmc.summary semantics (pure lookup):
    #   deployed  -> artifact has its own tmc_task
    #   worker    -> reachable via parent (not in orphaned set, not deployed)
    #   orphaned  -> listed in summary.orphaned_candidates
    header_row = r + 1
    headers = ["Artifact", "Type", "Deployed?", "Environments", "In prod?",
               "Classification", "Task IDs", "Provenance"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    orphan_set = set(orphaned)
    r = header_row + 1
    for a in sorted(doc["artifacts"], key=lambda x: (x["type"], x["name"])):
        task = a.get("tmc_task")
        deployed = task is not None
        if deployed:
            classification = "deployed"
        elif a["name"] in orphan_set:
            classification = "orphaned"
        else:
            classification = "worker / reachable"
        prov = (task or {}).get("provenance", "tmc") if deployed else "tmc"
        cells = [
            ws.cell(row=r, column=1, value=a["name"]),
            ws.cell(row=r, column=2, value=a["type"]),
            ws.cell(row=r, column=3, value="yes" if deployed else "no"),
            ws.cell(row=r, column=4, value=", ".join((task or {}).get("deployed_in_environments", []))),
            ws.cell(row=r, column=5, value="yes" if (task or {}).get("in_prod") else "no"),
            ws.cell(row=r, column=6, value=classification),
            ws.cell(row=r, column=7, value=", ".join((task or {}).get("task_ids", []))),
            ws.cell(row=r, column=8, value=prov),
        ]
        if deployed:
            _prov_fill(cells, prov)
        r += 1
    ws.auto_filter.ref = f"A{header_row}:H{max(header_row + 1, r - 1)}"
    _autosize(ws, [30, 12, 11, 22, 9, 18, 20, 12])


def sheet_dependencies(wb, doc: dict) -> None:
    """External jars/libs and the upgrade-risk (version-drift) view. Pure lookup:
    per-jar lib/version/provenance come from the already-computed per-artifact
    dependency records; pinned/drift come from the project-level dependencies block."""
    ws = wb.create_sheet("Dependencies")
    deps = doc.get("dependencies", {})
    distinct_jars = deps.get("distinct_jars", [])
    drift = deps.get("version_drift", [])
    if not distinct_jars and not drift:
        _header(ws, ["Jar", "Library", "Version", "Pinned?", "Used by", "Provenance"])
        ws.cell(row=2, column=1, value="(no external jar/library dependencies detected)")
        _autosize(ws, [30, 22, 14, 9, 40, 12])
        return

    # Index per-jar facts from per-artifact dependency records (already analyzed).
    pinned_set = set(deps.get("version_pinned_jars", []))
    jar_info: dict[str, dict] = {}
    jar_users: dict[str, set[str]] = {}
    for a in doc["artifacts"]:
        for d in a.get("dependencies", []):
            jar = d.get("jar")
            if not jar:
                continue
            jar_info.setdefault(jar, d)
            jar_users.setdefault(jar, set()).add(a["name"])

    _header(ws, ["Jar", "Library", "Version", "Pinned?", "Used by", "Provenance"])
    r = 2
    for jar in distinct_jars:
        info = jar_info.get(jar, {})
        prov = info.get("provenance", deps.get("provenance", "static"))
        cells = [
            ws.cell(row=r, column=1, value=jar),
            ws.cell(row=r, column=2, value=info.get("lib_base")),
            ws.cell(row=r, column=3, value=info.get("version")),
            ws.cell(row=r, column=4, value="yes" if jar in pinned_set else "no"),
            ws.cell(row=r, column=5, value=", ".join(sorted(jar_users.get(jar, [])))),
            ws.cell(row=r, column=6, value=prov),
        ]
        _prov_fill(cells, prov)
        r += 1
    ws.auto_filter.ref = f"A1:F{max(2, r - 1)}"

    # Version-drift section — the upgrade-risk view.
    r += 1
    ws.cell(row=r, column=1, value="Version drift (upgrade risk)").font = Font(bold=True, size=12)
    r += 1
    drift_header = r
    for c, h in enumerate(["Library", "Versions", "Jars"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    r += 1
    if not drift:
        ws.cell(row=r, column=1, value="(none — no library referenced in more than one version)")
        r += 1
    else:
        for d in drift:
            cells = [
                ws.cell(row=r, column=1, value=d.get("lib_base")),
                ws.cell(row=r, column=2, value=", ".join(d.get("versions", []))),
                ws.cell(row=r, column=3, value=", ".join(d.get("jars", []))),
            ]
            _prov_fill(cells, "manual")  # drift is an action item -> amber
            r += 1

    for flag in deps.get("upgrade_risk_flags", []):
        ws.cell(row=r, column=1, value="⚠ " + flag).font = Font(italic=True)
        r += 1
    _autosize(ws, [30, 22, 14, 9, 40, 12])


def sheet_gaps(wb, doc: dict) -> None:
    ws = wb.create_sheet("Gaps")
    _header(ws, ["Gap ID", "Kind", "Reference", "Description", "Suggested question",
                 "Resolution (fill in)", "Provenance"])
    r = 2
    for g in doc["gaps"]:
        ref = ", ".join(f"{k}={v}" for k, v in g["ref"].items() if v)
        cells = [
            ws.cell(row=r, column=1, value=g["gap_id"]),
            ws.cell(row=r, column=2, value=g["kind"]),
            ws.cell(row=r, column=3, value=ref),
            ws.cell(row=r, column=4, value=g["description"]),
            ws.cell(row=r, column=5, value=g["suggested_question"]),
            ws.cell(row=r, column=6, value=g.get("resolution") or ""),
            ws.cell(row=r, column=7, value=g["provenance"]),
        ]
        _prov_fill(cells, g["provenance"])
        r += 1
    ws.auto_filter.ref = f"A1:G{max(2, r - 1)}"
    _autosize(ws, [12, 22, 26, 46, 46, 30, 12])


def render(doc: dict, out_path: Path | str) -> Path:
    """Render the canonical intake document to an .xlsx workbook. Returns the path."""
    _require_openpyxl()
    wb = Workbook()
    sheet_summary(wb, doc)
    sheet_infrastructure(wb, doc)
    sheet_interfaces(wb, doc)
    sheet_artifacts(wb, doc)
    sheet_systems(wb, doc)
    sheet_system_usage(wb, doc)
    sheet_complexity(wb, doc)
    sheet_findings(wb, doc)
    sheet_dependencies(wb, doc)
    sheet_orchestration(wb, doc)
    sheet_deployment(wb, doc)
    sheet_gaps(wb, doc)
    out = Path(out_path)
    wb.save(out)
    return out


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Render intake JSON to an .xlsx workbook (phase 2).")
    p.add_argument("--in", dest="in_path", required=True, help="canonical intake JSON")
    p.add_argument("--out", dest="out_path", required=True, help="output .xlsx path")
    return p


def main(argv: Optional[list[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    doc = json.loads(Path(args.in_path).read_text(encoding="utf-8"))
    out = render(doc, args.out_path)
    print(f"wrote {out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
