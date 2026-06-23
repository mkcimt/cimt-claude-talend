"""Smoke + enrichment tests for the Excel renderer. Skipped when openpyxl is not
installed (the renderer is an optional, phase-2-only dependency).

Two regimes are exercised:
  * a static-only doc (tmc empty) — every sheet must still render, degrading to a
    placeholder where TMC/dependency data is absent;
  * an enriched doc (project_intake + tmc_intake via the FakeTmcClient pattern,
    plus jar dependencies with version drift) — the TMC/Deployment/Dependencies
    sheets must render real rows.

The renderer is PURE: it never analyzes, only lays out the canonical document, so
the enriched doc is produced by the real analyzers, not hand-built here.
"""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import project_intake as pi  # noqa: E402
import intake_to_excel as xl  # noqa: E402
import tmc_intake  # noqa: E402
from test_project_intake import build_fixture, _write, _node  # noqa: E402
from test_tmc_intake import FakeTmcClient  # noqa: E402

try:
    import openpyxl  # noqa: F401
    HAVE = True
except ImportError:
    HAVE = False


def _cells(ws, col=1):
    return {ws.cell(row=r, column=col).value for r in range(1, ws.max_row + 1)}


def _build_enriched(root: Path):
    """Real fixture + a job carrying the *same* driver jar pinned to two versions
    (version drift) + read-only TMC enrichment. Returns the canonical doc."""
    build_fixture(root)
    # A job that hard-references one library in two versions -> version drift,
    # the upgrade-risk signal the Dependencies sheet must surface. Generic names.
    _write(root, "process/deps", "j_with_libs", "0.1", "ProcessItem",
           nodes=_node("tDBInput", "db1", DRIVER_JAR="genericdriver_V1.jar")
                 + _node("tDBConnection", "db2", DRIVER_JAR="genericdriver-2.0.jar")
                 + _node("tLibraryLoad", "lib1", LIBRARY="custom-utils-1.2.3.jar"))
    doc = pi.analyze(root)
    tmc_intake.enrich(doc, client=FakeTmcClient(), generated_at="2026-06-23T00:00:00Z")
    return doc


@unittest.skipUnless(HAVE, "openpyxl not installed")
class TestExcelRenderer(unittest.TestCase):
    EXPECTED_SHEETS = ("Summary", "Infrastructure", "Interfaces", "Artifacts",
                       "Systems", "System Read-Write", "Complexity", "Findings",
                       "Dependencies", "Orchestration (TMC)", "Deployment", "Gaps")

    def test_static_only_doc_renders_all_sheets(self):
        """A doc with no TMC enrichment still produces every sheet, degrading to
        placeholders for the TMC/dependency views."""
        with tempfile.TemporaryDirectory() as d:
            root = Path(d)
            build_fixture(root)
            doc = pi.analyze(root)
            self.assertFalse(doc["tmc"].get("enriched"))
            out = root / "intake.xlsx"
            xl.render(doc, out)
            self.assertTrue(out.exists() and out.stat().st_size > 0)

            wb = openpyxl.load_workbook(out)
            for name in self.EXPECTED_SHEETS:
                self.assertIn(name, wb.sheetnames)

            # Artifacts sheet has a header row + one row per artifact.
            ws = wb["Artifacts"]
            self.assertEqual(ws.cell(row=1, column=1).value, "Name")
            self.assertEqual(ws.max_row - 1, len(doc["artifacts"]))

            # Systems sheet lists Oracle.
            self.assertIn("Oracle", _cells(wb["Systems"], col=2))

            # TMC-driven sheets degrade gracefully (placeholder, no crash).
            self.assertIn("(empty", str(wb["Orchestration (TMC)"].cell(2, 1).value))
            self.assertIn("(empty", str(wb["Deployment"].cell(2, 1).value))

            # Complexity gained the LLM columns.
            self.assertIn("Needs LLM review",
                          {wb["Complexity"].cell(1, c).value
                           for c in range(1, wb["Complexity"].max_column + 1)})

    def test_enriched_doc_populates_tmc_and_dependency_sheets(self):
        with tempfile.TemporaryDirectory() as d:
            root = Path(d)
            doc = _build_enriched(root)
            self.assertTrue(doc["tmc"]["enriched"])
            out = root / "intake_enriched.xlsx"
            xl.render(doc, out)
            wb = openpyxl.load_workbook(out)

            for name in self.EXPECTED_SHEETS:
                self.assertIn(name, wb.sheetnames)

            # --- Infrastructure: environments + engines + workspaces rendered. ---
            infra_cats = _cells(wb["Infrastructure"], col=1)
            self.assertIn("Environment", infra_cats)
            self.assertIn("Engine", infra_cats)
            self.assertIn("Workspace", infra_cats)
            env_names = _cells(wb["Infrastructure"], col=2)
            self.assertTrue({"prd", "dev"} & env_names)

            # --- Orchestration: the TMC plan is rendered (no placeholder). ---
            self.assertEqual(wb["Orchestration (TMC)"].cell(1, 1).value, "Plan")
            plan_names = _cells(wb["Orchestration (TMC)"], col=1)
            self.assertIn("plan1", plan_names)

            # --- Deployment: overview counts + per-artifact classification. ---
            dep = wb["Deployment"]
            overview_labels = _cells(dep, col=1)
            self.assertIn("Deployable total", overview_labels)
            self.assertIn("Reachable via parent (worker)", overview_labels)
            self.assertIn("Orphaned candidates", overview_labels)
            # the deployed job appears classified as deployed; orphan as orphaned.
            classifications = _cells(dep, col=6)
            self.assertIn("deployed", classifications)
            self.assertIn("orphaned", classifications)
            self.assertIn("worker / reachable", classifications)
            # per-env count from summary present somewhere in column 1/2.
            self.assertIn("Per environment", overview_labels)

            # --- Dependencies: distinct jars + version-drift section. ---
            deps_ws = wb["Dependencies"]
            self.assertEqual(deps_ws.cell(1, 1).value, "Jar")
            jar_col = _cells(deps_ws, col=1)
            self.assertIn("genericdriver_V1.jar", jar_col)
            self.assertIn("genericdriver-2.0.jar", jar_col)
            # the drift section header + the drifting library base.
            self.assertIn("Version drift (upgrade risk)", jar_col)
            self.assertIn("genericdriver", _cells(deps_ws, col=1))
            # pinned flag rendered for the version-pinned jars.
            self.assertIn("yes", _cells(deps_ws, col=4))

            # --- Findings: one row per artifacts[].findings entry, header intact. ---
            findings_ws = wb["Findings"]
            self.assertEqual(findings_ws.cell(1, 1).value, "Artifact")
            self.assertEqual(
                [findings_ws.cell(1, c).value for c in range(1, 7)],
                ["Artifact", "Severity", "Category", "Location", "Detail", "Provenance"],
            )
            expected_finding_rows = sum(len(a.get("findings", [])) for a in doc["artifacts"])
            if expected_finding_rows:
                self.assertEqual(findings_ws.max_row - 1, expected_finding_rows)
                self.assertTrue(_cells(findings_ws, col=2) - {"Artifact"})  # severities present

            # --- Summary: TMC + dependency overview present when enriched. ---
            summary_labels = _cells(wb["Summary"], col=1)
            self.assertIn("TMC (read-only)", summary_labels)
            self.assertIn("Deployed in prod", summary_labels)
            self.assertIn("Dependencies & upgrade risk", summary_labels)


if __name__ == "__main__":
    unittest.main(verbosity=2)
